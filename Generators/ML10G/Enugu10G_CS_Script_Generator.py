import openpyxl
import os
import fileinput
import sys


def Generate_Enugu_CS_Script(CP_ID, ui):

    # Open the first Excel file
    workbook1 = openpyxl.load_workbook('../../Config/ML_10GLLD/EnuguLLD/eptp10g.xlsx')
    worksheet1 = workbook1['eptp10g']

    # Open the second Excel file
    workbook2 = openpyxl.load_workbook('../../Config/ML_10GLLD/EnuguLLD/eslld10g.xlsx')
    worksheet2 = workbook2['eslld10g']

    # Open the third Excel file
    workbook3 = openpyxl.load_workbook('../../Config/ML_10GLLD/EnuguLLD/sysip2023.xlsx')
    worksheet3 = workbook3['sysip2023']

    created_files = []

    # Find the row numbers for SiteID name in the first file
    found_rows1 = []
    for row in worksheet1.iter_rows(min_row=2, min_col=1, values_only=True):
        if row[2] == CP_ID:  # made changes to be reviewed
            found_rows1.append(row)

    # Check if the SiteID name was found in the first file
    if not found_rows1:
        error_message = f"Could not find {CP_ID} in {worksheet1}."
        ui.showNotification(error_message)

    else:
        success_message = f"10G_CS Script has been Generated for {CP_ID} with required details from LLDs provided."
        ui.showNotification(success_message)

        # Get the details for the SiteID from the first file
        for index, row in enumerate(found_rows1, start=1):
            sites_details1 = row[1:17]

            # for rd_ip
            rd_ip = sites_details1[13]
            octets_rd_ip = rd_ip.split(".")
            for i in range(len(octets_rd_ip)):
                octets_rd_ip[i] = octets_rd_ip[i].zfill(3)
            removed_octets_rd_ip = octets_rd_ip[-2:]
            for i in range(len(removed_octets_rd_ip)):
                if len(removed_octets_rd_ip[i]) < 3:
                    removed_octets_rd_ip[i] = removed_octets_rd_ip[i].zfill(3)
            joined_octets = "".join(removed_octets_rd_ip)
            result_rd_ip = (joined_octets)

            # for rsvp_ip needs review
            rsvp_ip = sites_details1[13]
            octets = rsvp_ip.split(".")
            for i in range(len(octets)):
                octets[i] = octets[i].zfill(3)
            # Removing the third digit from the second octet
            removed_digit_oct1 = octets[1][0]
            # to get the 2nd & 3rd digit of 2nd octet
            removed_digit_oct2a = octets[1][1] + octets[1][2]
            # to get the 1st & 2nd digit of 3rd octet
            removed_digit_oct2 = octets[2][0] + octets[2][1]
            # to get the 4th digit of the 3rd octet + all digits in 4th octet
            removed_digit_oct3 = octets[3][0] + octets[3][1] + octets[3][2]
            # modifying the new 1st octet
            modified_first_octet = octets[0][:4] + removed_digit_oct1
            # modifying the new 2nd octet
            modified_second_octet = removed_digit_oct2a + octets[1][:-3] + removed_digit_oct2
            # modifying the new 3rd octet
            modified_third_octet = octets[2][2] + removed_digit_oct3
            # Creating the updated list of octets
            updated_octets = [f"{modified_first_octet}.{modified_second_octet}.{modified_third_octet}"]
            # Joining the digits tpogether
            for i in range(len(updated_octets)):
                if len(updated_octets[i]) < 4:
                    updated_octets[i] = updated_octets[i].zfill(4)
            joined_octets = ".".join(updated_octets)
            result_rsvp_ip = joined_octets

            # naming the file
            file_name = f"ModProj_{sites_details1[2]}_Enugu.txt"
            print(sites_details1[2])

            # Create a folder with the same name pattern as the file name
            folder_name = f"{CP_ID}_Enugu_ML6692"
            base_folder_path = "Enugu_Generated_Scripts"
            folder_path = os.path.join(os.getcwd(), base_folder_path, folder_name)
            os.makedirs(folder_path, exist_ok=True)

            # Specify the file path within the created folder
            file_path = os.path.join(folder_path, file_name)
            created_files.append(file_path)

            # Write site details to a text file for each found row in the first file
            with open(file_path, "w") as file:
                file.write("config\n")
                file.write(" no capability vrf-lite\n")
                file.write(f" router-id {sites_details1[13]}\n")
                file.write("!\n")
                file.write("!\n")
                file.write("ip vrf ran_enugu\n")
                file.write(f" router-id {sites_details1[13]}\n")
                file.write(f" rd 64909:{result_rd_ip}0300\n")
                file.write(" route-target import 64909:300\n")
                file.write(" route-target import 64909:1500\n")
                file.write(" route-target export 64909:300\n")
                file.write(" exit\n")
                file.write("!\n")
                file.write("!\n")
                file.write("ip vrf ran_oam_enugu\n")
                file.write(f" router-id {sites_details1[13]}\n")
                file.write(f" rd 64909:{result_rd_ip}1000\n")
                file.write(" route-target import 64999:6490003\n")
                file.write(" route-target export 64909:202\n")
                file.write(" exit\n")
                file.write("!\n")
                file.write("!\n")
                file.write("ip vrf lte_ran-gprs_gn\n")
                file.write(f" router-id {sites_details1[13]}\n")
                file.write(f" rd 64909:{result_rd_ip}0145\n")
                file.write(" route-target import 64909:1500\n")
                file.write(" route-target import 64999:145\n")
                file.write(" route-target export 64909:145\n")
                file.write(" exit\n")
                file.write("!\n")
                file.write("!\n")
                file.write("router rsvp\n")
                file.write(" keep-multiplier 3\n")
                file.write(" hello-interval 10\n")
                file.write(" explicit-null\n")
                file.write(" exit\n")
                file.write("!\n")
                file.write("!\n")
                file.write("router ldp\n")
                file.write(f" router-id {sites_details1[13]}\n")
                file.write(" control-mode ordered\n")
                file.write(f" transport-address ipv4 {sites_details1[13]}\n")
                file.write(" exit\n")
                file.write("!\n")
                file.write("!\n")
                file.write("interface ethernet 1/9/4\n")
                file.write(" lan\n")
                file.write("  speed full-duplex100\n")
                file.write("  no autoneg\n")
                file.write("  exit\n")
                file.write(" bridge-port\n")
                file.write("  l3enable 222\n")
                file.write("  exit\n")
                file.write(" exit\n")
                file.write("!\n")
                file.write("!\n")
                file.write("interface ethernet 1/9/5\n")
                file.write(" lan\n")
                file.write("  speed full-duplex100\n")
                file.write("  no autoneg\n")
                file.write("  exit\n")
                file.write(" bridge-port\n")
                file.write("  l3enable 333\n")
                file.write("  exit\n")
                file.write(" exit\n")
                file.write("!\n")
                file.write("!\n")
                file.write("interface ethernet 1/9/6\n")
                file.write(" bridge-port\n")
                file.write("  l3enable 441\n")
                file.write("  l3enable 444\n")
                file.write("  exit\n")
                file.write(" exit\n")
                file.write("!\n")
                file.write("!\n")
                file.write("interface ethernet 1/1/1\n")
                file.write(" bridge-port\n")
                file.write(f"  l3enable {sites_details1[11]}\n")
                file.write("  exit\n")
                file.write(" exit\n")
                file.write("!\n")
                file.write("!\n")

    # Find the row number for SiteID name in the second file
    found_rows2 = []
    for row in worksheet2.iter_rows(min_row=2, min_col=1, values_only=True):
        if row[2] == CP_ID:  # made changes to be reviewed
            found_rows2.append(row)

    # Check if the SiteID name was found in the first file
    if not found_rows2:
        error_message = f"Could not find {CP_ID} in {worksheet2}."
        ui.showNotification(error_message)

    else:
        success_message = f"10G_CS Script has been Generated for {CP_ID} with required details from LLDs provided."
        ui.showNotification(success_message)

        # Get the details for the SiteID from the first file
        for index, row in enumerate(found_rows2, start=1):
            sites_details2 = row[1:18]

            # naming the file
            file_name = f"ModProj_{sites_details2[2]}_Enugu.txt"
            print(sites_details2[2])

            # Create a folder with the same name pattern as the file name
            folder_name = f"{CP_ID}_Enugu_ML6692"
            base_folder_path = "Enugu_Generated_Scripts"
            folder_path = os.path.join(os.getcwd(), base_folder_path, folder_name)
            os.makedirs(folder_path, exist_ok=True)

            # Specify the file path within the created folder
            file_path = os.path.join(folder_path, file_name)
            created_files.append(file_path)

            # Write site details to the same text file for each found row in the second file
            with open(file_path, "a") as file:
                file.write("interface ip 1/9/4.222\n")
                file.write(" ip vrf forwarding ran_enugu\n")
                file.write(f" ip address {sites_details2[10]}/30\n")
                file.write(" no shutdown\n")
                file.write(" exit\n")
                file.write("!\n")
                file.write("!\n")
                file.write("interface ip 1/9/5.333\n")
                file.write(" ip vrf forwarding ran_enugu\n")
                file.write(f" ip address {sites_details2[13]}/30\n")
                file.write(" no shutdown\n")
                file.write(f" exit\n")
                file.write("!\n")
                file.write("!\n")
                file.write("interface ip 1/9/6.441\n")
                file.write(" ip vrf forwarding ran_oam_enugu\n")
                file.write(f" ip address {sites_details2[7]}/30\n")
                file.write(" no shutdown\n")
                file.write(" exit\n")
                file.write("!\n")
                file.write("!\n")
                file.write("interface ip 1/9/6.444\n")
                file.write(" ip vrf forwarding lte_ran-gprs_gn\n")
                file.write(f" ip address {sites_details2[5]}/30\n")
                file.write(" no shutdown\n")
                file.write(" exit\n")
                file.write("!\n")
                file.write("!\n")  # redefine a new function here

    # Find the row number for SiteID name in the second file
    found_rows3 = []
    for row in worksheet1.iter_rows(min_row=2, min_col=1, values_only=True):
        if row[2] == CP_ID:  # made changes to be reviewed
            found_rows3.append(row)

    # Check if the SiteID name was found in the first file
    if not found_rows3:
        error_message = f"Could not find {CP_ID} in {worksheet1}."
        ui.showNotification(error_message)

    else:
        success_message = f"10G_CS Script has been Generated for {CP_ID} with required details from LLDs provided."
        ui.showNotification(success_message)

        # Get the details for the SiteID from the first file
        for index, row in enumerate(found_rows3, start=1):
            sites_details1 = row[1:18]

            # for rsvp_ip needs review
            rsvp_ip = sites_details1[13]
            octets = rsvp_ip.split(".")
            for i in range(len(octets)):
                octets[i] = octets[i].zfill(3)
            # Removing the third digit from the second octet
            removed_digit_oct1 = octets[1][0]
            # to get the 2nd & 3rd digit of 2nd octet
            removed_digit_oct2a = octets[1][1] + octets[1][2]
            # to get the 1st & 2nd digit of 3rd octet
            removed_digit_oct2 = octets[2][0] + octets[2][1]
            # to get the 4th digit of the 3rd octet + all digits in 4th octet
            removed_digit_oct3 = octets[3][0] + octets[3][1] + octets[3][2]
            # modifying the new 1st octet
            modified_first_octet = octets[0][:4] + removed_digit_oct1
            # modifying the new 2nd octet
            modified_second_octet = removed_digit_oct2a + octets[1][:-3] + removed_digit_oct2
            # modifying the new 3rd octet
            modified_third_octet = octets[2][2] + removed_digit_oct3
            # Creating the updated list of octets
            updated_octets = [f"{modified_first_octet}.{modified_second_octet}.{modified_third_octet}"]
            # Joining the digits tpogether
            for i in range(len(updated_octets)):
                if len(updated_octets[i]) < 4:
                    updated_octets[i] = updated_octets[i].zfill(4)
            joined_octets = ".".join(updated_octets)
            result_rsvp_ip = joined_octets

            # naming the file
            file_name = f"ModProj_{sites_details1[2]}_Enugu.txt"

            # Create a folder with the same name pattern as the file name
            folder_name = f"{CP_ID}_Enugu_ML6692"
            base_folder_path = "Enugu_Generated_Scripts"
            folder_path = os.path.join(os.getcwd(), base_folder_path, folder_name)
            os.makedirs(folder_path, exist_ok=True)

            # Specify the file path within the created folder
            file_path = os.path.join(folder_path, file_name)
            created_files.append(file_path)

            # Write site details to the same text file for each found row in the second file
            with open(file_path, "a") as file:
                file.write(f"interface ip 1/1/1.{sites_details1[11]}\n")
                file.write(f" ip address {sites_details1[6]}/31\n")
                file.write(" no shutdown\n")
                file.write(" label-switching\n")
                file.write(" ip router isis isis25\n")
                file.write(" isis circuit-type level-1\n")
                file.write(" mpls ldp-igp sync isis level-1 holddown-timer 180\n")
                file.write(" isis ip bfd\n")
                file.write(" enable-ldp ipv4\n")
                file.write(" enable-rsvp\n")
                file.write(" rsvp keep-multiplier 3\n")
                file.write(" rsvp hello enable\n")
                file.write(" rsvp hello-interval 10\n")
                file.write(" rsvp hello-timeout 150\n")
                file.write(" bfd interval 100 minrx 100 multiplier 3\n")
                file.write(" exit\n")
                file.write("!\n")
                file.write("!\n")
                file.write("interface lo\n")
                file.write(f" ip address {sites_details1[13]}/32\n")
                file.write(" ip router isis isis25\n")
                file.write(" exit\n")
                file.write("!\n")
                file.write("!\n")
                file.write("interface ip lo.ran_enugu\n")
                file.write(f" ip address {sites_details1[13]}/32\n")
                file.write(" exit\n")
                file.write("!\n")
                file.write("!\n")
                file.write("interface ip lo.ran_oam_enugu\n")
                file.write(f" ip address {sites_details1[13]}/32\n")
                file.write(" exit\n")
                file.write("!\n")
                file.write("!\n")
                file.write("interface ip lo.lte_ran-gprs_gn\n")
                file.write(f" ip address {sites_details1[13]}/32\n")
                file.write(" exit\n")
                file.write("!\n")
                file.write("!\n")
                file.write("router isis isis25\n")
                file.write(" is-type level-1\n")
                file.write(" metric-style wide\n")
                file.write(" mpls traffic-eng level-1\n")
                file.write(f" mpls traffic-eng router-id {sites_details1[13]}\n")
                file.write(f" net 49.3025.{result_rsvp_ip}.00\n")  # needs review(add function)
                file.write(" exit\n")
                file.write("!\n")
                file.write("!\n")
                file.write("router bgp 64909\n")
                file.write(f" bgp router-id {sites_details1[13]}\n")

    # Find the row number for SiteID name in the third file
    found_rows4 = []
    for row in worksheet3.iter_rows(min_row=2, min_col=1, values_only=True):
        if row[2] == CP_ID:
            found_rows4.append(row)

    if not found_rows4:
        error_message = f"Could not find {CP_ID} in {worksheet3}."
        ui.showNotification(error_message)

    else:
        success_message = f"10G_CS Script has been Generated for {CP_ID} with required details from LLDs provided."
        ui.showNotification(success_message)

        # Get the details for the SiteID from the third file
        sites_details3 = found_rows4[0][1:18]  # Take the first match only

        appended_to_files = set()  # Keep track of files where information is appended
        for file_path in created_files:
            with open(file_path, "a") as file:
                if file_path not in appended_to_files:
                    file.write(f" neighbor {sites_details3[2]} remote-as 64909\n")
                    file.write(f" neighbor {sites_details3[2]} update-source lo\n")
                    file.write("!\n")
                    file.write("!\n")
                    file.write(" address-family vpnv4 unicast\n")
                    file.write(f" neighbor {sites_details3[2]} activate\n")
                    file.write(f" neighbor {sites_details3[2]} next-hop-self\n")
                    file.write(" exit\n")
                    file.write("!\n")
                    file.write("!\n")
                    file.write(" address-family ipv4 vrf ran_enugu\n")
                    file.write(" redistribute connected\n")
                    file.write(" redistribute static\n")
                    file.write(" exit\n")
                    file.write("!\n")
                    file.write("!\n")
                    file.write(" address-family ipv4 vrf ran_oam_enugu\n")
                    file.write(" redistribute connected\n")
                    file.write(" redistribute static\n")
                    file.write(" exit\n")
                    file.write("!\n")
                    file.write("!\n")
                    file.write(" address-family ipv4 vrf lte_ran-gprs_gn\n")
                    file.write(" redistribute connected\n")
                    file.write(" redistribute static\n")
                    file.write(" exit\n")
                    file.write("exit\n")
                    file.write("!\n")
                    file.write("!\n")  # redefine another function here
                    appended_to_files.add(file_path)  # Mark the file as appended

    # Find the row number for SiteID name in the second file
    found_rows5 = []
    for row in worksheet1.iter_rows(min_row=2, min_col=1, values_only=True):
        if row[2] == CP_ID:  # made changes to be reviewed
            found_rows5.append(row)

    # Check if the SiteID name was found in the first file
    if not found_rows5:
        error_message = f"Could not find {CP_ID} in {worksheet1}."
        ui.showNotification(error_message)

    else:
        success_message = f"10G_CS Script has been Generated for {CP_ID} with required details from LLDs provided."
        ui.showNotification(success_message)

        # Get the details for the SiteID from the first file
        for index, row in enumerate(found_rows5, start=1):
            sites_details1 = row[1:18]

            # naming the file
            file_name = f"ModProj_{sites_details1[2]}_Enugu.txt"

            # Create a folder with the same name pattern as the file name
            folder_name = f"{CP_ID}_Enugu_ML6692"
            base_folder_path = "Enugu_Generated_Scripts"
            folder_path = os.path.join(os.getcwd(), base_folder_path, folder_name)
            os.makedirs(folder_path, exist_ok=True)

            # Specify the file path within the created folder
            file_path = os.path.join(folder_path, file_name)
            created_files.append(file_path)

            with open(file_path, "a") as file:
                file.write(f"rsvp-path NFEvia{sites_details1[1]}\n")

    found_rows = []
    for row in worksheet1.iter_rows(min_row=2, min_col=1, values_only=True):
        if row[2] == CP_ID:
            found_rows.append(row)

    if not found_rows:
        error_message = f"Could not find {CP_ID} in {worksheet1}."
        ui.showNotification(error_message)

    else:
        success_message = f"10G_CS Script has been Generated for {CP_ID} with required details from LLDs provided."
        ui.showNotification(success_message)

        # Get the details for the SiteID from the third file
        sites_details1 = found_rows[0][1:18]  # Take the first match only

        appended_to_files = set()  # Keep track of files where information is appended
        for file_path in created_files:
            with open(file_path, "a") as file:
                if file_path not in appended_to_files:
                    file.write(f" {sites_details1[13]}\n")
                    appended_to_files.add(file_path)  # Mark the file as appended

    found_rows6 = []
    for row in worksheet3.iter_rows(min_row=2, min_col=1, values_only=True):
        if row[2] == CP_ID:
            found_rows6.append(row)

    if not found_rows6:
        error_message = f"Could not find {CP_ID} in {worksheet3}."
        ui.showNotification(error_message)

    else:
        success_message = f"10G_CS Script has been Generated for {CP_ID} with required details from LLDs provided."
        ui.showNotification(success_message)

        # Get the details for the SiteID from the third file
        sites_details3 = found_rows6[0][1:18]  # Take the first match only

        appended_to_files = set()  # Keep track of files where information is appended
        for file_path in created_files:
            with open(file_path, "a") as file:
                if file_path not in appended_to_files:
                    file.write(f" {sites_details3[2]} strict\n")
                    file.write(" exit\n")
                    file.write("!\n")
                    file.write("!\n")
                    appended_to_files.add(file_path)  # Mark the file as appended

    # Find the row number for SiteID name in the second file
    found_rows7 = []
    for row in worksheet1.iter_rows(min_row=2, min_col=1, values_only=True):
        if row[2] == CP_ID:  # made changes to be reviewed
            found_rows7.append(row)

    # Check if the SiteID name was found in the first file
    if not found_rows7:
        error_message = f"Could not find {CP_ID} in {worksheet1}."
        ui.showNotification(error_message)

    else:
        success_message = f"10G_CS Script has been Generated for {CP_ID} with required details from LLDs provided."
        ui.showNotification(success_message)

        # Get the details for the SiteID from the first file
        for index, row in enumerate(found_rows7, start=1):
            sites_details1 = row[1:18]

            # naming the file
            file_name = f"ModProj_{sites_details1[2]}_Enugu.txt"

            # Create a folder with the same name pattern as the file name
            folder_name = f"{CP_ID}_Enugu_ML6692"
            base_folder_path = "Enugu_Generated_Scripts"
            folder_path = os.path.join(os.getcwd(), base_folder_path, folder_name)
            os.makedirs(folder_path, exist_ok=True)

            # Specify the file path within the created folder
            file_path = os.path.join(folder_path, file_name)
            created_files.append(file_path)

            with open(file_path, "a") as file:
                file.write(f"rsvp-trunk NFEvia{sites_details1[1]} ipv4\n")
                file.write(f" primary path NFEvia{sites_details1[1]}\n")
                file.write(f" from {sites_details1[6]}\n")

    # Find the row number for SiteID name in the third file
    found_rows8 = []
    for row in worksheet3.iter_rows(min_row=2, min_col=1, values_only=True):
        if row[2] == CP_ID:
            found_rows8.append(row)

    if not found_rows8:
        error_message = f"Could not find {CP_ID} in {worksheet3}."
        ui.showNotification(error_message)

    else:
        success_message = f"10G_CS Script has been Generated for {CP_ID} with required details from LLDs provided."
        ui.showNotification(success_message)

        # Get the details for the SiteID from the third file
        sites_details3 = found_rows8[0][1:18]  # Take the first match only

        appended_to_files = set()  # Keep track of files where information is appended
        for file_path in created_files:
            with open(file_path, "a") as file:
                if file_path not in appended_to_files:
                    file.write(f" to {sites_details3[2]}\n")
                    file.write(" exit\n")
                    file.write("!\n")
                    file.write("!\n")
                    appended_to_files.add(file_path)  # Mark the file as appended
                    
            # return file_name, file_name
